import pandas as pd
import re
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



dupl_comp_check = {
    "HZ_IMP_PARTIES_T":["Party Original System Reference", "*Party Type", "Party Number", "Organization Name"],
    "HZ_IMP_PARTYSITES_T":["Party Site Original System Reference", "Location Original System Reference", "Party Site Name", "Party Site Number"]
}
# Define expected data formats (using cleaned column names)
EXPECTED_FORMATS = {
    "CUSTOMER_TYPE": ["Organization", "Person"],
    "COUNTRY": ["US", "IN", "UK", "CA", "AU"],  # Example list of valid countries
}

# Regular expressions for validating email and phone number
EMAIL_REGEX = r"(^[\w\.\-]+@[\w\-]+\.[\w\-\.]+$)"
PHONE_REGEX = r"^\+?[\d\s\-()]{7,}$"  # Simple regex: allows optional '+' and various separators

def fetch_data_info(file_path_name, st_name):
    wb = load_workbook(file_path_name, data_only=True)
    if st_name != None:
        ws = wb[st_name]
        header_comments = {}
        for cell in ws[4]:  # First row is typically the header
            if cell.comment:
                comment_text = cell.comment.text.strip()
                
                
                # Split comment text by lines
                comment_lines = comment_text.split('\n')
                

                # Parse field blocks and handle edge cases
                try:
                    values = parse_field_blocks(comment_lines)
                    header_comments[cell.value] = values
                except IndexError as e:
                    print(f"IndexError: {e} when processing comment for column '{cell.value}'")
                    continue  # Skip this column if there's an error

        #print(header_comments)

        # Append to the results
    
    
    return header_comments

def parse_field_blocks(lines):
    fields = []
    i = 0
    n = len(lines)

    while i < n:
        # Skip blank lines
        while i < n and lines[i].strip() == "":
            i += 1
        if i >= n:
            break

        line = lines[i].strip()

        # Case 1: Inline column name and data type e.g. "PARTY_ORIG_SYSTEM VARCHAR2(30)"
        inline_match = re.match(r"^(\w+)\s+(\w+)\s*\(?\s*(\d*)\s*\)?$", line)
        if inline_match:
            column_name = inline_match.group(1)
            data_type = inline_match.group(2)
            length = inline_match.group(3) if inline_match.group(3) else None
            i += 1
            # Skip empty lines
            while i < n and lines[i].strip() == "":
                i += 1
            # Remaining lines = comment
            comment_lines = []
            while i < n and lines[i].strip() != "":
                comment_lines.append(lines[i].strip())
                i += 1
            fields.append({
                "internal_name": column_name,
                "data_type": data_type,
                "length": length,
                "constraint": None,
                "comment": " ".join(comment_lines)
            })
            continue

        # Case 2: Multi-line: name, optional constraint, data_type, comment
        column_name = line
        i += 1

        # Skip empty lines
        while i < n and lines[i].strip() == "":
            i += 1

        constraint = None
        data_type = None
        length = None

        # Constraint (optional)
        if i < n and re.match(r"(?i)^NOT NULL$", lines[i].strip()):
            constraint = lines[i].strip().upper()
            i += 1
            while i < n and lines[i].strip() == "":
                i += 1

        # Data type and optional length
        if i < n:
            dtype_line = lines[i].strip()
            match = re.match(r"(\w+)\s*\(?\s*(\d*)\s*\)?", dtype_line)
            if match:
                data_type = match.group(1)
                length = match.group(2) if match.group(2) else None
            i += 1

        # Skip empty lines
        while i < n and lines[i].strip() == "":
            i += 1

        # Comment (all remaining lines)
        comment_lines = []
        while i < n and lines[i].strip() != "":
            comment_lines.append(lines[i].strip())
            i += 1

        fields.append({
            "internal_name": column_name,
            "data_type": data_type,
            "length": length,
            "constraint": constraint,
            "comment": " ".join(comment_lines)
        })

    return fields
        
    #print(df)

def space_tab_newline_check(df):
    spc_df = fetch_leading_trailing_whitespace(df)
    issue_df = spc_df.dropna(how='all')
    return issue_df


def fetch_leading_trailing_whitespace(df):
    whitespace_info = {}
    for col in df.columns:
        whitespace_info[col] = df[col].apply(lambda x: x if isinstance(x, str) and (x.startswith(' ') or x.endswith(' ') or '\n' in x) else None)
    
    return pd.DataFrame(whitespace_info)

def fetch_column_lov(file_path,st_name):
    wb = load_workbook(file_path, data_only=True)
    final_dict = {}
    header_row = 4  # Headers are in row 3
    lov_list = []
    #lov_dict = {}  # Dictionary to store LOVs for each shee
    if st_name != None:
        ws = wb[st_name]
        headers = {
            get_column_letter(col_idx): ws.cell(row=header_row, column=col_idx).value
            for col_idx in range(1, ws.max_column + 1)
        }
        single_sheet_lovs = {}
        for dv in ws.data_validations.dataValidation:
            if dv.formula1 and dv.type == "list":  # Check if it's a drop-down list
                first_cell_ref = str(next(iter(dv.sqref))).split(":")[0]
                print(first_cell_ref)
                col_id = re.findall("[a-zA-Z]+",first_cell_ref,0)
                #print(col_id)
                #first_cell = list(dv.sqref.cells)[0]  # Get first cell (row, col)
                #col_letter = get_column_letter(first_cell[1])  # Convert column index to lette

                for y in col_id:
                    column_name = headers.get(y)  # Get column name from headers
                    print(column_name)
                #print(column_name)
                values = dv.formula1.strip('"').split(",")  # Extract LOV values
                print(values)
            
                single_sheet_lovs[column_name] = values

        #if single_sheet_lovs:
         #   lov_dict[sheet_name]=sheet_lovs  # Store LOVs for this sheet
        print(single_sheet_lovs)
        final_dict = single_sheet_lovs
    else:
        sheets_to_process = wb.sheetnames[1:]
# Iterate through all sheets in the workbook
        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            lov_dict = {}
    # Read the headers from row 3 for this sheet
            headers = {
                get_column_letter(col_idx): ws.cell(row=header_row, column=col_idx).value
                for col_idx in range(1, ws.max_column + 1)
                }
        
    # Extract LOVs from Data Validation Rules
            sheet_lovs = {}

            for dv in ws.data_validations.dataValidation:
                if dv.formula1 and dv.type == "list":  # Check if it's a drop-down list
                    first_cell_ref = str(next(iter(dv.sqref))).split(":")[0]
                    print(first_cell_ref)
                    col_id = re.findall("[a-zA-Z]+",first_cell_ref,0)
                #print(col_id)
                #first_cell = list(dv.sqref.cells)[0]  # Get first cell (row, col)
                #col_letter = get_column_letter(first_cell[1])  # Convert column index to lette

                    for y in col_id:
                        column_name = headers.get(y)  # Get column name from headers
                #print(column_name)
                        values = dv.formula1.strip('"').split(",")  # Extract LOV values
            
                        sheet_lovs[column_name] = values

                if sheet_lovs:
                    lov_dict[sheet_name]=sheet_lovs  # Store LOVs for this sheet
    
            lov_list.append(lov_dict)
        final_dict[file_path] = lov_list    
    #final_json = json.dumps(final_dict, indent=4)
    return final_dict

# Check for missing values in mandatory fields
def check_completeness(df, mandatory_columns):
    # Ensure we only check columns that exist in the DataFrame
    cols_to_check = [col for col in mandatory_columns if col in df.columns]
    missing_values = df[cols_to_check].isnull().sum()
    return missing_values[missing_values > 0]

# Check data consistency (e.g., format validation)
def check_consistency(df):
    inconsistent = {}
    # Validate expected formats for specific columns
    for col, valid_values in EXPECTED_FORMATS.items():
        if col in df.columns:
            invalid_rows = df[~df[col].isin(valid_values)]
            if not invalid_rows.empty:
                inconsistent[col] = invalid_rows[col].unique().tolist()

    # Validate email format if EMAIL column exists
    if "Email Address" in df.columns:
        invalid_emails = df[~df["Email Address"].fillna("").str.match(EMAIL_REGEX, na=False)]
        if not invalid_emails.empty:
            inconsistent["EMAIL"] = invalid_emails["Email Address"].unique().tolist()

    # Validate phone number format if PHONE_NUMBER column exists
    if "Phone Number" in df.columns:
        invalid_phones = df[~df["Phone Number"].fillna("").str.match(PHONE_REGEX, na=False)]
        if not invalid_phones.empty:
            inconsistent["Phone Number"] = invalid_phones["Phone Number"].unique().tolist()

    return inconsistent

# Check uniqueness (e.g., duplicate customer records)
def check_uniqueness(df,sht_name):
    dupl_records = df
    try:
        composite_columns = dupl_comp_check[sht_name]
    except Exception as e:
        composite_columns = None
    if composite_columns != None:
        dupl_records = df[df.duplicated(subset=composite_columns, keep=False)]
    else:
        dupl_records = []
    return dupl_records

# Check referential integrity (e.g., for organizations, ensure PARTY_NUMBER exists)
def check_integrity(df):
    orphan_accounts = df[(df["CUSTOMER_TYPE"] == "Organization") & (~df["PARTY_NUMBER"].notnull())]
    return orphan_accounts

# Generate a Data Quality Report
def generate_report(*input):
    final_report = {}
    if isinstance(input[0], pd.DataFrame):
        sht_name = input[1]
        print(sht_name)
        df =  input[0]   
        cols = []
        cols = df.columns
        mandatory_columns = [x for x in cols if (x.startswith("*") or x.endswith("*"))]
        
        completeness_issues = check_completeness(df, mandatory_columns)
        consistency_issues = check_consistency(df)
        duplicate_records = check_uniqueness(df,sht_name)
            #orphan_accounts = check_integrity(df)

        # Calculate a simple data quality score.
        missing_issue_count = completeness_issues.sum() if not completeness_issues.empty else 0
        inconsistency_issue_count = sum(len(v) for v in consistency_issues.values())
        duplicate_issue_count = len(duplicate_records)
            #orphan_issue_count = len(orphan_accounts)
        total_issues = missing_issue_count + inconsistency_issue_count + duplicate_issue_count 
            #+ orphan_issue_count
        lovs = fetch_column_lov(input[2], sht_name)
        # Quality score: penalize issues relative to total records.

        data_types = fetch_data_info(input[2], sht_name)
        combined_data_types = pd.concat([pd.DataFrame(rows).replace({"": None}).assign(section=section) 
                                 for section, rows in data_types.items()],ignore_index=True)
        final_data_types = combined_data_types[combined_data_types['data_type'].notna()]
        
        space_issues = space_tab_newline_check(df)

        quality_score = max(0, 100 - (total_issues / (len(df) + 1) * 100))  # Adding 1 to avoid division by zero
            #report[sht_name] = f"{quality_score:.2f}%"
        # Temporary fix, change the code with the real check
        dup_rec = {}
        comp_rec = {}
        if len(duplicate_records) > 0:
            dup_rec = duplicate_records.to_dict()

        if len(completeness_issues) > 0:
            comp_rec = completeness_issues.to_dict()

        report = {
                "Number of Rows": len(df),
                "Number of Columns": len(df.columns),
                "Mandatory Columns": mandatory_columns,
                "Missing Values": comp_rec,
                "Inconsistent Data": consistency_issues,
                "Duplicate Records Count": len(duplicate_records),
                "Duplicate Record Details": dup_rec, # returns dataframe
                "List of Values":lovs,
                "Data Types":final_data_types,
                "White Space Issues":space_issues,
                #"Orphan Accounts Count": len(orphan_accounts),
                "Total Records": len(df),
                "Quality Score": quality_score
            }
        final_report=report
    else:

    #process_sheets = read_fbdi(file_path)
        fbdi_sheets_mand_cols = []
        print(input[0])
        file = input[0]
        sheets = file.sheet_names[1:]
        print(len(sheets))
        
        x = 0
        while x < len(sheets):
            list_mand = []
            #print(len(sheets))
            df = pd.read_excel(input[0], sheet_name=sheets[x],header=3)
            sht_name = sheets[x]
            x = x+1
            columns = []
            columns = df.columns
            mandatory_columns = [x for x in columns if x.startswith("*")]
        
            completeness_issues = check_completeness(df, mandatory_columns)
            consistency_issues = check_consistency(df)
            duplicate_records = check_uniqueness(df,sht_name)
            #orphan_accounts = check_integrity(df)

        # Calculate a simple data quality score.
            missing_issue_count = completeness_issues.sum() if not completeness_issues.empty else 0
            inconsistency_issue_count = sum(len(v) for v in consistency_issues.values())
            duplicate_issue_count = len(duplicate_records)
            #orphan_issue_count = len(orphan_accounts)
            total_issues = missing_issue_count + inconsistency_issue_count + duplicate_issue_count 
            #+ orphan_issue_count

        # Quality score: penalize issues relative to total records.
            quality_score = max(0, 100 - (total_issues / (len(df) + 1) * 100))  # Adding 1 to avoid division by zero
            #report[sht_name] = f"{quality_score:.2f}%"
            report = {
                #"NUmber Of Rows":len(df),
                #"Number of Columns": len(df.columns),
                #"Mandatory Columns": mandatory_columns,
                #"Missing Values": completeness_issues.to_dict(),
                #"Inconsistent Data": consistency_issues,
                #"Duplicate Records Count": len(duplicate_records),
                "Duplicate_Record_Details":duplicate_records,
                #"Orphan Accounts Count": len(orphan_accounts),
                #"Total Records": len(df),
                "Quality_Score": quality_score
            }
            freport = []
            freport.append(report)
            final_report[sht_name]=freport
    #print(final_report)
    return final_report

# Example usage:
#fbdi_file_path = "/Users/srikarbala/Downloads/CustomerImportTemplate.xlsm"  # Replace with your actual file path
#data_quality_report = generate_report(fbdi_file_path)

# Print the report
#for key, value in data_quality_report.items():
#    print(f"{key}: {value}")